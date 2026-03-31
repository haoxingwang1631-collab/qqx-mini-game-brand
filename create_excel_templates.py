from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os

BASE = "/Users/haoxingwang/WorkBuddy/QQx小游戏品牌方案/02_配套Excel模板"

# Common styles
HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='1A73E8')
SUB_HEADER_FILL = PatternFill('solid', fgColor='4A90D9')
TITLE_FONT = Font(name='Arial', bold=True, size=14, color='1A73E8')
NORMAL_FONT = Font(name='Arial', size=10)
BOLD_FONT = Font(name='Arial', bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_header_row(ws, row, cols, fill=None):
    f = fill or HEADER_FILL
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = f
        cell.alignment = CENTER
        cell.border = THIN_BORDER

def style_data_area(ws, start_row, end_row, cols):
    for r in range(start_row, end_row+1):
        for c in range(1, cols+1):
            cell = ws.cell(row=r, column=c)
            cell.font = NORMAL_FONT
            cell.alignment = LEFT_WRAP
            cell.border = THIN_BORDER

def auto_width(ws, cols, min_w=12, max_w=35):
    for c in range(1, cols+1):
        ws.column_dimensions[get_column_letter(c)].width = min(max_w, max(min_w, 15))

def add_title(ws, title, cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 35

# ====== 1. 行业情报周报模板 ======
def create_intel_weekly():
    wb = Workbook()
    
    # Sheet1: 情报登记表
    ws = wb.active
    ws.title = "情报登记表"
    add_title(ws, "📊 行业情报周报 — 情报登记表", 10)
    headers = ["日期", "信息来源", "案例/资讯名称", "涉及游戏", "涉及媒体", "游戏品类", "营销玩法类型", "关键洞察", "与QQ关联度", "备注/链接"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    style_header_row(ws, 2, 10)
    style_data_area(ws, 3, 22, 10)
    
    dv_category = DataValidation(type="list", formula1='"休闲,SLG,模拟经营,卡牌,RPG,竞技,益智,社交,其他"')
    dv_play = DataValidation(type="list", formula1='"品牌联动,效果广告,达人合作,IP联名,互动挑战赛,开屏广告,信息流,直播推广,其他"')
    dv_relevance = DataValidation(type="list", formula1='"高,中,低"')
    ws.add_data_validation(dv_category)
    ws.add_data_validation(dv_play)
    ws.add_data_validation(dv_relevance)
    for r in range(3, 23):
        dv_category.add(ws.cell(row=r, column=6))
        dv_play.add(ws.cell(row=r, column=7))
        dv_relevance.add(ws.cell(row=r, column=9))
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 14
    ws.column_dimensions['J'].width = 30
    
    # Sheet2: 趋势汇总看板
    ws2 = wb.create_sheet("趋势汇总看板")
    add_title(ws2, "📈 趋势汇总看板", 5)
    
    sections = [
        ("品类维度趋势", ["品类", "本周动态数量", "环比变化", "趋势判断", "备注"]),
        ("媒体维度趋势", ["媒体平台", "本周动态数量", "环比变化", "趋势判断", "备注"]),
        ("玩法维度趋势", ["营销玩法", "本周出现次数", "环比变化", "趋势判断", "备注"])
    ]
    row = 2
    for section_title, cols in sections:
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws2.cell(row=row, column=1, value=f"▎{section_title}").font = BOLD_FONT
        row += 1
        for i, h in enumerate(cols, 1):
            ws2.cell(row=row, column=i, value=h)
        style_header_row(ws2, row, 5, SUB_HEADER_FILL)
        row += 1
        style_data_area(ws2, row, row+4, 5)
        row += 6
    auto_width(ws2, 5)
    
    # Sheet3: 重点案例深度分析
    ws3 = wb.create_sheet("重点案例深度分析")
    add_title(ws3, "🔍 重点案例深度分析", 6)
    headers3 = ["案例名称", "涉及游戏/媒体", "营销模式概述", "亮点分析", "对QQ的启发", "参考链接"]
    for i, h in enumerate(headers3, 1):
        ws3.cell(row=2, column=i, value=h)
    style_header_row(ws3, 2, 6)
    style_data_area(ws3, 3, 7, 6)
    for c in range(1, 7):
        ws3.column_dimensions[get_column_letter(c)].width = 25
    ws3.column_dimensions['C'].width = 35
    ws3.column_dimensions['D'].width = 35
    ws3.column_dimensions['E'].width = 35
    
    wb.save(os.path.join(BASE, "行业情报周报模板.xlsx"))
    print("✅ 行业情报周报模板.xlsx")

# ====== 2. 高潜客户筛选漏斗表 ======
def create_customer_funnel():
    wb = Workbook()
    
    # Sheet1: 客户基础信息池
    ws = wb.active
    ws.title = "客户基础信息池"
    add_title(ws, "🌊 客户基础信息池", 12)
    headers = ["序号", "客户/公司名称", "游戏名称", "游戏品类", "上线日期", "DAU(万)", "月流水(万)", "变现模式", "年轻用户占比", "内容调性", "已知投放平台", "信息来源"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    style_header_row(ws, 2, 12)
    style_data_area(ws, 3, 22, 12)
    
    dv_cat = DataValidation(type="list", formula1='"休闲,SLG,模拟经营,卡牌,RPG,竞技,益智,社交,其他"')
    dv_mode = DataValidation(type="list", formula1='"纯IAA,IAA+IAP混合,纯IAP,订阅制,其他"')
    dv_tone = DataValidation(type="list", formula1='"年轻潮流/社交向,泛娱乐,偏成熟/商务,其他"')
    ws.add_data_validation(dv_cat)
    ws.add_data_validation(dv_mode)
    ws.add_data_validation(dv_tone)
    for r in range(3, 23):
        dv_cat.add(ws.cell(row=r, column=4))
        dv_mode.add(ws.cell(row=r, column=8))
        dv_tone.add(ws.cell(row=r, column=10))
    auto_width(ws, 12, 13)
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    
    # Sheet2: 筛选评分表
    ws2 = wb.create_sheet("筛选评分表")
    add_title(ws2, "🎯 四维筛选评分表（自动加权计算）", 14)
    headers2 = ["客户名称", "游戏名称",
                "基础面-DAU", "基础面-上线时长", "基础面-品类契合", "基础面小计(25%)",
                "商业化-IAA收入", "商业化-投放活跃", "商业化-变现模式", "商业化小计(30%)",
                "QQ契合-画像重合", "QQ契合-调性匹配", "QQ契合-IP社交", "QQ契合小计(25%)"]
    # need more columns
    extra_headers = ["预算-历史预算", "预算-品牌意愿", "预算-竞媒活跃", "预算小计(20%)", "综合得分", "客户等级"]
    all_headers = headers2 + extra_headers
    for i, h in enumerate(all_headers, 1):
        ws2.cell(row=2, column=i, value=h)
    style_header_row(ws2, 2, len(all_headers))
    
    # Sub-dimension grouping colors
    fill_basic = PatternFill('solid', fgColor='E8F0FE')
    fill_biz = PatternFill('solid', fgColor='FFF2CC')
    fill_qq = PatternFill('solid', fgColor='D9EAD3')
    fill_budget = PatternFill('solid', fgColor='FCE5CD')
    fill_result = PatternFill('solid', fgColor='F4CCCC')
    
    for r in range(3, 23):
        for c in range(3, 7):
            ws2.cell(row=r, column=c).fill = fill_basic
        for c in range(7, 11):
            ws2.cell(row=r, column=c).fill = fill_biz
        for c in range(11, 15):
            ws2.cell(row=r, column=c).fill = fill_qq
        for c in range(15, 19):
            ws2.cell(row=r, column=c).fill = fill_budget
        for c in range(19, 21):
            ws2.cell(row=r, column=c).fill = fill_result
        
        # Subtotal formulas
        ws2.cell(row=r, column=6).value = f'=AVERAGE(C{r}:E{r})'  # 基础面小计
        ws2.cell(row=r, column=10).value = f'=AVERAGE(G{r}:I{r})'  # 商业化小计
        ws2.cell(row=r, column=14).value = f'=AVERAGE(K{r}:M{r})'  # QQ契合小计
        ws2.cell(row=r, column=18).value = f'=AVERAGE(O{r}:Q{r})'  # 预算小计
        ws2.cell(row=r, column=19).value = f'=F{r}*0.25+J{r}*0.3+N{r}*0.25+R{r}*0.2'  # 综合得分
        ws2.cell(row=r, column=20).value = f'=IF(S{r}>=4,"S",IF(S{r}>=3,"A",IF(S{r}>=2,"B","C")))'  # 等级
    
    style_data_area(ws2, 3, 22, len(all_headers))
    for c in range(1, len(all_headers)+1):
        ws2.column_dimensions[get_column_letter(c)].width = 14
    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 16
    
    # Add scoring guide
    ws2.cell(row=24, column=1, value="📋 评分标准：1-5分制").font = BOLD_FONT
    guide = [
        ["5分", "表现优秀（如DAU≥200万、月投放≥100万、年轻用户≥50%）"],
        ["3分", "表现中等（如DAU 50-200万、月投放30-100万、年轻用户30-40%）"],
        ["1分", "表现较弱（如DAU<50万、无投放记录、年轻用户<30%）"]
    ]
    for i, (score, desc) in enumerate(guide):
        ws2.cell(row=25+i, column=1, value=score).font = BOLD_FONT
        ws2.cell(row=25+i, column=2, value=desc)
    
    # Sheet3: 客户分档结果
    ws3 = wb.create_sheet("客户分档结果")
    add_title(ws3, "📊 客户分档结果", 7)
    headers3 = ["客户名称", "游戏名称", "综合得分", "客户等级", "建议策略", "跟进节奏", "销售确认"]
    for i, h in enumerate(headers3, 1):
        ws3.cell(row=2, column=i, value=h)
    style_header_row(ws3, 2, 7)
    style_data_area(ws3, 3, 22, 7)
    auto_width(ws3, 7, 15)
    
    # Sheet4: 筛选日志
    ws4 = wb.create_sheet("筛选日志")
    add_title(ws4, "📝 筛选操作日志", 5)
    headers4 = ["日期", "操作人", "操作内容", "新增客户数", "备注"]
    for i, h in enumerate(headers4, 1):
        ws4.cell(row=2, column=i, value=h)
    style_header_row(ws4, 2, 5)
    style_data_area(ws4, 3, 22, 5)
    auto_width(ws4, 5)
    
    wb.save(os.path.join(BASE, "高潜客户筛选漏斗表.xlsx"))
    print("✅ 高潜客户筛选漏斗表.xlsx")

# ====== 3. 竞媒策略分析框架表 ======
def create_competitor_analysis():
    wb = Workbook()
    
    # Sheet1: 竞媒基础信息
    ws = wb.active
    ws.title = "竞媒基础信息"
    add_title(ws, "⚔️ 竞媒基础信息对比", 7)
    headers = ["分析维度", "抖音", "小红书", "快手", "支付宝", "QQ", "差异化洞察"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    style_header_row(ws, 2, 7)
    
    dimensions = [
        "平台定位", "小游戏DAU/MAU", "核心用户年龄段", "用户性别比例",
        "城市线级分布", "小游戏入口位置", "流量分配机制", "广告产品线",
        "合作模式", "开屏广告价格区间", "信息流CPM区间", "互动广告形式",
        "品牌专区/定制", "达人合作体系", "IP联名案例", "最低合作门槛",
        "结算方式", "数据报告能力", "核心优势", "核心劣势"
    ]
    for i, dim in enumerate(dimensions, 3):
        ws.cell(row=i, column=1, value=dim).font = BOLD_FONT
    
    # Highlight QQ column
    qq_fill = PatternFill('solid', fgColor='E8F0FE')
    for r in range(3, 3+len(dimensions)):
        ws.cell(row=r, column=6).fill = qq_fill
    
    style_data_area(ws, 3, 2+len(dimensions), 7)
    ws.column_dimensions['A'].width = 18
    for c in range(2, 8):
        ws.column_dimensions[get_column_letter(c)].width = 22

    # Sheet2: 案例拆解表
    ws2 = wb.create_sheet("案例拆解表")
    add_title(ws2, "🔍 竞媒案例拆解", 9)
    headers2 = ["媒体平台", "合作游戏", "游戏品类", "营销形式", "创意玩法描述", "效果数据", "亮点分析", "可借鉴点", "适用于QQ的改造建议"]
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=2, column=i, value=h)
    style_header_row(ws2, 2, 9)
    style_data_area(ws2, 3, 22, 9)
    
    dv_platform = DataValidation(type="list", formula1='"抖音,小红书,快手,支付宝,其他"')
    ws2.add_data_validation(dv_platform)
    for r in range(3, 23):
        dv_platform.add(ws2.cell(row=r, column=1))
    for c in range(1, 10):
        ws2.column_dimensions[get_column_letter(c)].width = 20
    ws2.column_dimensions['E'].width = 30
    ws2.column_dimensions['H'].width = 25
    ws2.column_dimensions['I'].width = 30
    
    # Sheet3: QQ差异化优势对照
    ws3 = wb.create_sheet("QQ差异化优势对照")
    add_title(ws3, "💎 QQ差异化优势对照表", 5)
    headers3 = ["差异化维度", "QQ优势描述", "竞品对比（劣势所在）", "可转化为卖点的话术", "适用客户品类"]
    for i, h in enumerate(headers3, 1):
        ws3.cell(row=2, column=i, value=h)
    style_header_row(ws3, 2, 5)
    
    advantages = [
        "年轻用户浓度", "社交关系链深度", "QQ空间/看点流量",
        "二次元/潮流文化基因", "游戏基因与联动能力", "差异化场景（如QQ音乐、QQ阅读）",
        "品效协同能力", "数据能力与精准度"
    ]
    for i, adv in enumerate(advantages, 3):
        ws3.cell(row=i, column=1, value=adv).font = BOLD_FONT
    style_data_area(ws3, 3, 2+len(advantages), 5)
    for c in range(1, 6):
        ws3.column_dimensions[get_column_letter(c)].width = 28
    
    wb.save(os.path.join(BASE, "竞媒策略分析框架表.xlsx"))
    print("✅ 竞媒策略分析框架表.xlsx")

# ====== 4. 商机管理追踪表 ======
def create_opportunity_tracker():
    wb = Workbook()
    
    # Sheet1: 商机主表
    ws = wb.active
    ws.title = "商机主表"
    add_title(ws, "💼 商机管理追踪表 — 主表", 20)
    
    # Group headers (row 2)
    groups = [
        ("基础信息", 1, 8, PatternFill('solid', fgColor='1A73E8')),
        ("商机状态", 9, 14, PatternFill('solid', fgColor='0B8043')),
        ("会议纪要提取", 15, 21, PatternFill('solid', fgColor='E37400')),
        ("结果", 22, 25, PatternFill('solid', fgColor='C5221F')),
    ]
    
    headers = [
        # 基础信息 1-8
        "商机ID", "客户名称", "游戏名称", "游戏品类", "客户联系人", "联系方式", "所属销售", "商机来源",
        # 商机状态 9-14
        "商机阶段", "意向评级", "预估预算(万)", "目标档期", "签约概率", "预估签约金额(万)",
        # 会议纪要提取 15-21
        "最近会议日期", "参会人员", "客户核心关注点", "明确需求清单", "客户顾虑/异议", "竞品对比提及", "下一步To-Do",
        # 结果 22-25
        "最终签约金额(万)", "合作档期", "流失原因", "流失详细说明"
    ]
    
    # Group headers
    for title, start, end, fill in groups:
        ws.merge_cells(start_row=2, start_column=start, end_row=2, end_column=end)
        cell = ws.cell(row=2, column=start, value=title)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = CENTER
        for c in range(start, end+1):
            ws.cell(row=2, column=c).fill = fill
            ws.cell(row=2, column=c).border = THIN_BORDER
    
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(headers))
    style_data_area(ws, 4, 23, len(headers))
    
    # Data validations
    dv_cat = DataValidation(type="list", formula1='"休闲,SLG,模拟经营,卡牌,RPG,竞技,益智,社交,其他"')
    dv_stage = DataValidation(type="list", formula1='"初步接触,需求确认,方案提报,商务谈判,签约,执行中,已结项,流失"')
    dv_grade = DataValidation(type="list", formula1='"S,A,B,C"')
    dv_source = DataValidation(type="list", formula1='"主动挖掘,情报发现,客户咨询,老客续投,内部推荐,其他"')
    dv_loss = DataValidation(type="list", formula1='"预算不足,转投竞品,需求不匹配,内部调整,时间冲突,其他"')
    
    for dv in [dv_cat, dv_stage, dv_grade, dv_source, dv_loss]:
        ws.add_data_validation(dv)
    
    for r in range(4, 24):
        dv_cat.add(ws.cell(row=r, column=4))
        dv_stage.add(ws.cell(row=r, column=9))
        dv_grade.add(ws.cell(row=r, column=10))
        dv_source.add(ws.cell(row=r, column=8))
        dv_loss.add(ws.cell(row=r, column=24))
    
    for c in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['Q'].width = 25
    ws.column_dimensions['R'].width = 25
    ws.column_dimensions['S'].width = 25
    ws.column_dimensions['U'].width = 25
    
    # Sheet2: 跟进记录日志
    ws2 = wb.create_sheet("跟进记录日志")
    add_title(ws2, "📝 跟进记录日志", 7)
    headers2 = ["商机ID", "客户名称", "跟进日期", "跟进方式", "跟进内容摘要", "阶段变更", "下一步计划"]
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=2, column=i, value=h)
    style_header_row(ws2, 2, 7)
    style_data_area(ws2, 3, 32, 7)
    
    dv_method = DataValidation(type="list", formula1='"腾讯会议,电话,企微消息,邮件,线下拜访,其他"')
    ws2.add_data_validation(dv_method)
    for r in range(3, 33):
        dv_method.add(ws2.cell(row=r, column=4))
    auto_width(ws2, 7, 16)
    ws2.column_dimensions['E'].width = 35
    ws2.column_dimensions['G'].width = 25
    
    # Sheet3: 漏斗转化统计
    ws3 = wb.create_sheet("漏斗转化统计")
    add_title(ws3, "📊 商机漏斗转化统计", 4)
    headers3 = ["商机阶段", "当前数量", "转化率", "平均停留天数"]
    for i, h in enumerate(headers3, 1):
        ws3.cell(row=2, column=i, value=h)
    style_header_row(ws3, 2, 4)
    
    stages = ["初步接触", "需求确认", "方案提报", "商务谈判", "签约", "执行中", "已结项", "流失"]
    for i, stage in enumerate(stages, 3):
        ws3.cell(row=i, column=1, value=stage).font = BOLD_FONT
    style_data_area(ws3, 3, 10, 4)
    auto_width(ws3, 4, 18)
    
    wb.save(os.path.join(BASE, "商机管理追踪表.xlsx"))
    print("✅ 商机管理追踪表.xlsx")

# ====== 5. 投后复盘与ROI追踪表 ======
def create_roi_tracker():
    wb = Workbook()
    
    # Sheet1: 投放数据登记
    ws = wb.active
    ws.title = "投放数据登记"
    add_title(ws, "📊 投后复盘与ROI追踪 — 数据登记", 16)
    headers = [
        "项目编号", "客户名称", "游戏名称", "品类", "投放档期", "资源位",
        "总预算(万)", "实际花费(万)", "总曝光量(万)", "总点击量(万)", "点击率CTR",
        "转化量", "转化率CVR", "CPA(元)", "ROI", "备注"
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    style_header_row(ws, 2, 16)
    
    for r in range(3, 17):
        ws.cell(row=r, column=11).value = f'=IF(I{r}=0,"-",J{r}/I{r})'
        ws.cell(row=r, column=11).number_format = '0.00%'
        ws.cell(row=r, column=13).value = f'=IF(J{r}=0,"-",L{r}/J{r})'
        ws.cell(row=r, column=13).number_format = '0.00%'
        ws.cell(row=r, column=14).value = f'=IF(L{r}=0,"-",H{r}*10000/L{r})'
        ws.cell(row=r, column=14).number_format = '#,##0.00'
        ws.cell(row=r, column=15).value = f'=IF(H{r}=0,"-",L{r}/(H{r}*10000))'
        ws.cell(row=r, column=15).number_format = '0.00%'
    
    style_data_area(ws, 3, 16, 16)
    for c in range(1, 17):
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 16
    
    # Sheet2: 效果对比分析
    ws2 = wb.create_sheet("效果对比分析")
    add_title(ws2, "📈 效果对比分析", 7)
    headers2 = ["指标", "预估值", "实际值", "达成率", "行业基准", "vs基准", "分析说明"]
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=2, column=i, value=h)
    style_header_row(ws2, 2, 7)
    
    metrics = ["总曝光量(万)", "总点击量(万)", "CTR", "转化量", "CVR", "CPA(元)", "ROI", "总花费(万)"]
    for i, m in enumerate(metrics, 3):
        ws2.cell(row=i, column=1, value=m).font = BOLD_FONT
        ws2.cell(row=i, column=4).value = f'=IF(B{i}=0,"-",C{i}/B{i})'
        ws2.cell(row=i, column=4).number_format = '0.0%'
        ws2.cell(row=i, column=6).value = f'=IF(E{i}=0,"-",C{i}/E{i})'
        ws2.cell(row=i, column=6).number_format = '0.0%'
    style_data_area(ws2, 3, 10, 7)
    auto_width(ws2, 7, 16)
    
    # Sheet3: 优化建议记录
    ws3 = wb.create_sheet("优化建议记录")
    add_title(ws3, "💡 优化建议记录", 5)
    headers3 = ["项目/客户", "问题发现", "优化建议", "优先级", "执行状态"]
    for i, h in enumerate(headers3, 1):
        ws3.cell(row=2, column=i, value=h)
    style_header_row(ws3, 2, 5)
    style_data_area(ws3, 3, 17, 5)
    
    dv_priority = DataValidation(type="list", formula1='"高,中,低"')
    dv_status = DataValidation(type="list", formula1='"待执行,执行中,已完成,已取消"')
    ws3.add_data_validation(dv_priority)
    ws3.add_data_validation(dv_status)
    for r in range(3, 18):
        dv_priority.add(ws3.cell(row=r, column=4))
        dv_status.add(ws3.cell(row=r, column=5))
    auto_width(ws3, 5, 20)
    ws3.column_dimensions['B'].width = 30
    ws3.column_dimensions['C'].width = 30
    
    wb.save(os.path.join(BASE, "投后复盘与ROI追踪表.xlsx"))
    print("✅ 投后复盘与ROI追踪表.xlsx")

# ====== 6. 客户分层运营表 ======
def create_customer_tier():
    wb = Workbook()
    
    # Sheet1: 客户价值评估
    ws = wb.active
    ws.title = "客户价值评估"
    add_title(ws, "👥 客户分层运营 — 价值评估", 12)
    headers = [
        "客户名称", "游戏名称", "累计投放金额(万)", "投放次数", "平均ROI",
        "最近投放日期", "续投意愿", "客户满意度", "综合评分", "客户等级",
        "续投策略", "触达频率"
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    style_header_row(ws, 2, 12)
    style_data_area(ws, 3, 22, 12)
    
    dv_will = DataValidation(type="list", formula1='"强,中,弱,未知"')
    dv_satis = DataValidation(type="list", formula1='"非常满意,满意,一般,不满意"')
    dv_level = DataValidation(type="list", formula1='"S-战略客户,A-优质客户,B-培育客户,C-观望客户"')
    dv_freq = DataValidation(type="list", formula1='"每周,双周,月度,季度"')
    ws.add_data_validation(dv_will)
    ws.add_data_validation(dv_satis)
    ws.add_data_validation(dv_level)
    ws.add_data_validation(dv_freq)
    for r in range(3, 23):
        dv_will.add(ws.cell(row=r, column=7))
        dv_satis.add(ws.cell(row=r, column=8))
        dv_level.add(ws.cell(row=r, column=10))
        dv_freq.add(ws.cell(row=r, column=12))
    auto_width(ws, 12, 15)
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['K'].width = 25
    
    # Sheet2: 续投策略与触达计划
    ws2 = wb.create_sheet("续投策略与触达计划")
    add_title(ws2, "📅 续投策略与触达计划", 8)
    headers2 = ["客户名称", "客户等级", "下次触达日期", "触达方式", "触达内容/话术", "续投方案准备状态", "预估续投金额(万)", "备注"]
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=2, column=i, value=h)
    style_header_row(ws2, 2, 8)
    style_data_area(ws2, 3, 22, 8)
    
    dv_prep = DataValidation(type="list", formula1='"待准备,准备中,已就绪,已提交"')
    ws2.add_data_validation(dv_prep)
    for r in range(3, 23):
        dv_prep.add(ws2.cell(row=r, column=6))
    auto_width(ws2, 8, 16)
    ws2.column_dimensions['E'].width = 30
    
    # Sheet3: 客户生命周期状态
    ws3 = wb.create_sheet("客户生命周期")
    add_title(ws3, "🔄 客户生命周期状态", 7)
    headers3 = ["客户名称", "首次合作日期", "最近合作日期", "生命周期阶段", "累计合作次数", "生命周期价值LTV(万)", "状态变更记录"]
    for i, h in enumerate(headers3, 1):
        ws3.cell(row=2, column=i, value=h)
    style_header_row(ws3, 2, 7)
    style_data_area(ws3, 3, 22, 7)
    
    dv_stage = DataValidation(type="list", formula1='"新客,活跃客户,沉默客户,流失预警,已流失,已挽回"')
    ws3.add_data_validation(dv_stage)
    for r in range(3, 23):
        dv_stage.add(ws3.cell(row=r, column=4))
    auto_width(ws3, 7, 18)
    
    wb.save(os.path.join(BASE, "客户分层运营表.xlsx"))
    print("✅ 客户分层运营表.xlsx")

# ====== 7. 标杆案例登记表 ======
def create_case_library():
    wb = Workbook()
    
    ws = wb.active
    ws.title = "标杆案例库"
    add_title(ws, "🏆 标杆案例登记表", 13)
    headers = [
        "案例编号", "客户名称", "游戏名称", "游戏品类", "合作档期",
        "投放预算(万)", "资源组合", "创意玩法描述", "效果数据摘要",
        "亮点总结", "可复用要素", "素材存档路径", "录入日期"
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    style_header_row(ws, 2, 13)
    style_data_area(ws, 3, 17, 13)
    
    dv_cat = DataValidation(type="list", formula1='"休闲,SLG,模拟经营,卡牌,RPG,竞技,益智,社交,其他"')
    ws.add_data_validation(dv_cat)
    for r in range(3, 18):
        dv_cat.add(ws.cell(row=r, column=4))
    
    for c in range(1, 14):
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 25
    ws.column_dimensions['J'].width = 25
    ws.column_dimensions['K'].width = 25
    ws.column_dimensions['L'].width = 25
    
    # Sheet2: 案例标签索引
    ws2 = wb.create_sheet("案例标签索引")
    add_title(ws2, "🏷️ 案例标签索引（快速检索）", 5)
    headers2 = ["标签维度", "标签值", "关联案例编号", "案例数量", "备注"]
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=2, column=i, value=h)
    style_header_row(ws2, 2, 5)
    
    tags = ["按品类", "按资源类型", "按预算量级", "按创意玩法", "按效果表现"]
    for i, tag in enumerate(tags, 3):
        ws2.cell(row=i, column=1, value=tag).font = BOLD_FONT
    style_data_area(ws2, 3, 7, 5)
    auto_width(ws2, 5, 18)
    
    wb.save(os.path.join(BASE, "标杆案例登记表.xlsx"))
    print("✅ 标杆案例登记表.xlsx")


# ====== Execute All ======
if __name__ == "__main__":
    create_intel_weekly()
    create_customer_funnel()
    create_competitor_analysis()
    create_opportunity_tracker()
    create_roi_tracker()
    create_customer_tier()
    create_case_library()
    print("\n🎉 所有7份Excel模板创建完成！")
